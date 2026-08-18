import io
import os
from datetime import date, datetime, timedelta
from typing import Dict, Any, List
from sqlalchemy.orm import Session
from sqlalchemy import func

# ReportLab imports
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

# openpyxl imports
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# matplotlib imports
import matplotlib
matplotlib.use('Agg')  # Non-interactive backend
import matplotlib.pyplot as plt

from app.db.models import Product, Inventory, Sale, SaleItem
from app.services.forecasting import calculate_forecast_for_product
from app.services.advisor_service import collect_business_evidence, generate_deterministic_action_plan
from app.services.local_intelligence_service import generate_grounded_local_intelligence

def collect_report_data(db: Session, shop_id: str = "shop_001", period: str = "7d") -> Dict[str, Any]:
    """
    CENTRALIZED REPORT DATA PIPELINE.
    Single source of truth for PDF, Excel, and PNG reports.
    CRITICAL GUARANTEE: Reads data only. Zero database mutations.
    """
    today = date.today()
    if period == "today":
        days = 1
        start_date = today
    elif period == "30d":
        days = 30
        start_date = today - timedelta(days=29)
    else:  # Default "7d"
        days = 7
        start_date = today - timedelta(days=6)

    # 1. Fetch Sales in Date Window
    sales = db.query(Sale).filter(
        Sale.shop_id == shop_id,
        func.date(Sale.created_at) >= start_date,
        func.date(Sale.created_at) <= today
    ).all()

    rev = round(sum(float(s.total_amount) for s in sales), 2)
    cost = round(sum(float(s.total_cost) for s in sales), 2)
    profit = round(rev - cost, 2)
    margin = round((profit / rev * 100), 2) if rev > 0 else 0.0

    # 2. Fetch Products & Inventory
    products = db.query(Product).all()
    inventories = db.query(Inventory).filter(Inventory.shop_id == shop_id).all()
    inv_map = {i.product_id: i.quantity for i in inventories}
    inv_value = round(sum(inv.quantity * float(p.purchase_price) for inv in inventories for p in products if p.id == inv.product_id), 2)

    healthy_cnt = 0
    low_cnt = 0
    out_cnt = 0
    for p in products:
        q = inv_map.get(p.id, 0)
        if q == 0:
            out_cnt += 1
        elif q <= p.reorder_level:
            low_cnt += 1
        else:
            healthy_cnt += 1

    # 3. Product Sales Performance
    perf_query = db.query(
        SaleItem.product_id,
        func.sum(SaleItem.quantity).label('units_sold'),
        func.sum(SaleItem.quantity * SaleItem.unit_price).label('revenue'),
        func.sum(SaleItem.profit).label('profit')
    ).join(Sale, Sale.id == SaleItem.sale_id)\
     .filter(Sale.shop_id == shop_id, func.date(Sale.created_at) >= start_date)\
     .group_by(SaleItem.product_id).all()

    perf_map = {r.product_id: r for r in perf_query}

    top_sellers = []
    profit_leaders = []
    slow_moving = []

    for p in products:
        rec = perf_map.get(p.id)
        u_sold = int(rec.units_sold) if rec else 0
        p_rev = round(float(rec.revenue), 2) if rec else 0.0
        p_prof = round(float(rec.profit), 2) if rec else 0.0
        qty = inv_map.get(p.id, 0)

        item = {
            "product_id": p.id,
            "name": p.name,
            "category": p.category,
            "unit": p.unit,
            "purchase_price": float(p.purchase_price),
            "selling_price": float(p.selling_price),
            "stock_quantity": qty,
            "units_sold": u_sold,
            "revenue": p_rev,
            "profit": p_prof
        }

        if u_sold > 0:
            top_sellers.append(item)
            if p_prof > 0:
                profit_leaders.append(item)
        elif qty > 0:
            slow_moving.append(item)

    top_sellers.sort(key=lambda x: x["units_sold"], reverse=True)
    profit_leaders.sort(key=lambda x: x["profit"], reverse=True)

    # 4. Forecast Data
    reorder_items = []
    for p in products:
        qty = inv_map.get(p.id, 0)
        fc = calculate_forecast_for_product({}, qty, p.reorder_level, 7, today)
        rec_pur = fc["planning_suggestion"]["recommended_purchase"]
        if fc["stock_status"] in ["OUT_OF_STOCK", "LOW_STOCK", "AT_RISK"] or rec_pur > 0:
            reorder_items.append({
                "product_id": p.id,
                "name": p.name,
                "current_stock": qty,
                "reorder_level": p.reorder_level,
                "forecast_demand": fc["forecast_daily_demand"],
                "days_of_stock": fc["days_of_stock"],
                "stock_status": fc["stock_status"],
                "recommended_purchase": rec_pur
            })

    # 5. Advisor Plan
    evidence = collect_business_evidence(db, shop_id)
    advisor_plan = generate_deterministic_action_plan(evidence)

    # 6. Seasonal & Local Intelligence
    local_intel = generate_grounded_local_intelligence(db, shop_id, today=today)

    return {
        "metadata": {
            "shop_id": shop_id,
            "shop_name": "Sri Lakshmi General Store",
            "locality": "Ameerpet, Hyderabad",
            "period_code": period,
            "period_start": start_date.strftime('%Y-%m-%d'),
            "period_end": today.strftime('%Y-%m-%d'),
            "generated_at": datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC'),
            "data_source": "MARUTHI Verified Business Data"
        },
        "financials": {
            "revenue": rev,
            "cogs": cost,
            "profit": profit,
            "margin": margin,
            "inventory_value": inv_value
        },
        "inventory_health": {
            "total_products": len(products),
            "healthy_count": healthy_cnt,
            "low_stock_count": low_cnt,
            "out_of_stock_count": out_cnt,
            "inventory_value": inv_value
        },
        "sales_performance": {
            "total_sales_count": len(sales),
            "top_sellers": top_sellers[:5],
            "profit_leaders": profit_leaders[:5],
            "slow_moving": slow_moving[:5]
        },
        "forecast": {
            "reorder_items": reorder_items[:5],
            "at_risk_count": len(reorder_items)
        },
        "advisor": {
            "mode": advisor_plan.mode,
            "recommendations": [r.dict() for r in advisor_plan.recommendations]
        },
        "seasonal_local": {
            "current_season": local_intel.current_season,
            "upcoming_festivals": local_intel.upcoming_festivals,
            "location_name": local_intel.resolved_location_name,
            "recommendations": [r.dict() for r in local_intel.recommendations]
        }
    }

def generate_pdf_report(data: Dict[str, Any]) -> bytes:
    """
    Generate PDF Report using ReportLab.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, leading=22, textColor=colors.HexColor('#0f172a'))
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#475569'))
    h2_style = ParagraphStyle('H2', parent=styles['Heading2'], fontSize=12, leading=15, textColor=colors.HexColor('#1e293b'), spaceBefore=8, spaceAfter=4)
    normal_style = ParagraphStyle('NormalText', parent=styles['Normal'], fontSize=8.5, leading=11, textColor=colors.HexColor('#334155'))
    bold_style = ParagraphStyle('BoldText', parent=styles['Normal'], fontSize=8.5, leading=11, fontName='Helvetica-Bold', textColor=colors.HexColor('#0f172a'))

    elements = []
    meta = data["metadata"]
    fin = data["financials"]

    # Header
    elements.append(Paragraph("<b>MARUTHI — AI Retail Copilot</b>", title_style))
    elements.append(Paragraph(f"Shop: {meta['shop_name']} ({meta['locality']}) | Period: {meta['period_start']} to {meta['period_end']} ({meta['period_code']})", subtitle_style))
    elements.append(Spacer(1, 8))
    elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#cbd5e1'), spaceBefore=0, spaceAfter=8))

    # Business Summary Table
    elements.append(Paragraph("BUSINESS SUMMARY", h2_style))
    sum_table_data = [
        [Paragraph("<b>Metric</b>", bold_style), Paragraph("<b>Value</b>", bold_style)],
        [Paragraph("Revenue", normal_style), Paragraph(f"Rs. {fin['revenue']:,.2f}", normal_style)],
        [Paragraph("Cost of Goods Sold (COGS)", normal_style), Paragraph(f"Rs. {fin['cogs']:,.2f}", normal_style)],
        [Paragraph("Gross Profit", normal_style), Paragraph(f"Rs. {fin['profit']:,.2f}", normal_style)],
        [Paragraph("Gross Margin %", normal_style), Paragraph(f"{fin['margin']}%", normal_style)],
        [Paragraph("Total Inventory Valuation", normal_style), Paragraph(f"Rs. {fin['inventory_value']:,.2f}", normal_style)],
    ]
    t1 = Table(sum_table_data, colWidths=[240, 240])
    t1.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f1f5f9')),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
        ('PADDING', (0,0), (-1,-1), 4),
    ]))
    elements.append(t1)
    elements.append(Spacer(1, 10))

    # Inventory & Sales Health
    elements.append(Paragraph("INVENTORY & SALES PERFORMANCE", h2_style))
    inv_health = data["inventory_health"]
    inv_data = [
        [Paragraph("Total Products", normal_style), Paragraph(str(inv_health["total_products"]), normal_style)],
        [Paragraph("Healthy Stock Items", normal_style), Paragraph(str(inv_health["healthy_count"]), normal_style)],
        [Paragraph("Low Stock Items", normal_style), Paragraph(str(inv_health["low_stock_count"]), normal_style)],
        [Paragraph("Out of Stock Items", normal_style), Paragraph(str(inv_health["out_of_stock_count"]), normal_style)],
    ]
    t2 = Table(inv_data, colWidths=[240, 240])
    t2.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
        ('PADDING', (0,0), (-1,-1), 4),
    ]))
    elements.append(t2)
    elements.append(Spacer(1, 10))

    # Top Sellers
    elements.append(Paragraph("TOP SELLERS & PROFIT LEADERS", h2_style))
    top_sellers = data["sales_performance"]["top_sellers"]
    top_rows = [[Paragraph("<b>Product Name</b>", bold_style), Paragraph("<b>Units Sold</b>", bold_style), Paragraph("<b>Revenue</b>", bold_style), Paragraph("<b>Profit</b>", bold_style)]]
    for ts in top_sellers:
        top_rows.append([
            Paragraph(ts["name"], normal_style),
            Paragraph(str(ts["units_sold"]), normal_style),
            Paragraph(f"Rs. {ts['revenue']:,.2f}", normal_style),
            Paragraph(f"Rs. {ts['profit']:,.2f}", normal_style),
        ])
    if len(top_rows) == 1:
        top_rows.append([Paragraph("No sales recorded in period", normal_style), Paragraph("-", normal_style), Paragraph("-", normal_style), Paragraph("-", normal_style)])
    t3 = Table(top_rows, colWidths=[200, 80, 100, 100])
    t3.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f1f5f9')),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
        ('PADDING', (0,0), (-1,-1), 4),
    ]))
    elements.append(t3)
    elements.append(Spacer(1, 10))

    # MARUTHI Advisor & Seasonal Intelligence
    elements.append(Paragraph("MARUTHI ADVISOR & SEASONAL INTELLIGENCE", h2_style))
    adv = data["advisor"]["recommendations"]
    adv_text = adv[0]["recommendation_summary"] if adv else "Inventory healthy."
    sec_intel = data["seasonal_local"]
    intel_text = f"Season: {sec_intel['current_season']} | Festivals: {', '.join(sec_intel['upcoming_festivals'])}"

    adv_data = [
        [Paragraph("Tomorrow Priority Action", bold_style), Paragraph(adv_text, normal_style)],
        [Paragraph("Seasonal Drivers", bold_style), Paragraph(intel_text, normal_style)],
    ]
    t4 = Table(adv_data, colWidths=[160, 320])
    t4.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
        ('PADDING', (0,0), (-1,-1), 4),
    ]))
    elements.append(t4)
    elements.append(Spacer(1, 14))

    # Footer
    elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor('#cbd5e1'), spaceBefore=0, spaceAfter=6))
    elements.append(Paragraph(f"Generated at: {meta['generated_at']} | Source: {meta['data_source']} | Verified 100% Read-Only", subtitle_style))

    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()

def generate_xlsx_report(data: Dict[str, Any]) -> bytes:
    """
    Generate Excel XLSX Report using openpyxl.
    7 Distinct Sheets: Summary, Sales, Inventory, Product Performance, Forecast, Advisor, Seasonal & Local Intelligence.
    """
    wb = openpyxl.Workbook()
    
    # Styles
    font_header = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    font_title = Font(name='Calibri', size=14, bold=True, color='0F172A')
    font_bold = Font(name='Calibri', size=11, bold=True)
    fill_header = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
    border_thin = Border(left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'), top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1'))

    meta = data["metadata"]
    fin = data["financials"]

    # 1. Summary Sheet
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.freeze_panes = 'A5'

    ws1.cell(row=1, column=1, value="MARUTHI - AI Retail Copilot Business Report").font = font_title
    ws1.cell(row=2, column=1, value=f"Shop: {meta['shop_name']} ({meta['locality']}) | Period: {meta['period_start']} to {meta['period_end']}")
    ws1.cell(row=3, column=1, value=f"Generated: {meta['generated_at']} | Source: {meta['data_source']}")

    headers1 = ["Metric", "Value"]
    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=4, column=col, value=h)
        cell.font = font_header
        cell.fill = fill_header

    metrics_rows = [
        ("Revenue", fin["revenue"], '₹#,##0.00'),
        ("Cost of Goods Sold (COGS)", fin["cogs"], '₹#,##0.00'),
        ("Gross Profit", fin["profit"], '₹#,##0.00'),
        ("Gross Margin %", fin["margin"] / 100.0, '0.0%'),
        ("Total Inventory Valuation", fin["inventory_value"], '₹#,##0.00'),
        ("Total Sales Count", data["sales_performance"]["total_sales_count"], '#,##0'),
        ("Low Stock Items Count", data["inventory_health"]["low_stock_count"], '#,##0'),
        ("Out of Stock Items Count", data["inventory_health"]["out_of_stock_count"], '#,##0'),
    ]

    for r_idx, (m, v, fmt) in enumerate(metrics_rows, 5):
        c1 = ws1.cell(row=r_idx, column=1, value=m)
        c2 = ws1.cell(row=r_idx, column=2, value=v)
        c2.number_format = fmt
        c1.border = border_thin
        c2.border = border_thin

    # 2. Sales Sheet
    ws2 = wb.create_sheet(title="Sales")
    ws2.freeze_panes = 'A2'
    ws2.append(["Product Name", "Units Sold", "Revenue", "Profit"])
    for cell in ws2[1]:
        cell.font = font_header
        cell.fill = fill_header

    for ts in data["sales_performance"]["top_sellers"]:
        ws2.append([ts["name"], ts["units_sold"], ts["revenue"], ts["profit"]])

    # 3. Inventory Sheet
    ws3 = wb.create_sheet(title="Inventory")
    ws3.freeze_panes = 'A2'
    ws3.append(["Metric", "Count / Value"])
    for cell in ws3[1]:
        cell.font = font_header
        cell.fill = fill_header

    inv_h = data["inventory_health"]
    ws3.append(["Total Products", inv_h["total_products"]])
    ws3.append(["Healthy Stock Count", inv_h["healthy_count"]])
    ws3.append(["Low Stock Count", inv_h["low_stock_count"]])
    ws3.append(["Out of Stock Count", inv_h["out_of_stock_count"]])
    ws3.append(["Inventory Valuation", inv_h["inventory_value"]])

    # 4. Product Performance Sheet
    ws4 = wb.create_sheet(title="Product Performance")
    ws4.freeze_panes = 'A2'
    ws4.append(["Product Name", "Category", "Stock Qty", "Purchase Price", "Selling Price", "Units Sold", "Revenue", "Profit"])
    for cell in ws4[1]:
        cell.font = font_header
        cell.fill = fill_header

    for item in data["sales_performance"]["top_sellers"] + data["sales_performance"]["slow_moving"]:
        ws4.append([item["name"], item["category"], item["stock_quantity"], item["purchase_price"], item["selling_price"], item["units_sold"], item["revenue"], item["profit"]])

    # 5. Forecast Sheet
    ws5 = wb.create_sheet(title="Forecast")
    ws5.freeze_panes = 'A2'
    ws5.append(["Product Name", "Current Stock", "Reorder Level", "Forecast Demand (units/day)", "Days of Stock", "Stock Status", "Recommended Purchase"])
    for cell in ws5[1]:
        cell.font = font_header
        cell.fill = fill_header

    for fc in data["forecast"]["reorder_items"]:
        ws5.append([fc["name"], fc["current_stock"], fc["reorder_level"], fc["forecast_demand"], fc["days_of_stock"], fc["stock_status"], fc["recommended_purchase"]])

    # 6. Advisor Sheet
    ws6 = wb.create_sheet(title="Advisor")
    ws6.freeze_panes = 'A2'
    ws6.append(["Priority", "Category", "Title", "Recommendation Summary"])
    for cell in ws6[1]:
        cell.font = font_header
        cell.fill = fill_header

    for rec in data["advisor"]["recommendations"]:
        ws6.append([rec.get("priority", 1), rec.get("category", ""), rec.get("title", ""), rec.get("recommendation_summary", "")])

    # 7. Seasonal & Local Intelligence Sheet
    ws7 = wb.create_sheet(title="Seasonal & Local Intelligence")
    ws7.freeze_panes = 'A2'
    ws7.append(["Category", "Title", "Recommendation Summary", "Why Reason"])
    for cell in ws7[1]:
        cell.font = font_header
        cell.fill = fill_header

    for rec in data["seasonal_local"]["recommendations"]:
        ws7.append([rec.get("category", ""), rec.get("title", ""), rec.get("recommendation_summary", ""), rec.get("why_reason", "")])

    # Auto-adjust column widths across all sheets
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def generate_png_report(data: Dict[str, Any]) -> bytes:
    """
    Generate PNG Executive Summary Snapshot using matplotlib.
    Optimized for WhatsApp / mobile sharing.
    """
    fig, ax = plt.subplots(figsize=(10.8, 13.5), dpi=100)  # 1080x1350 mobile card ratio
    fig.patch.set_facecolor('#0f172a')
    ax.set_facecolor('#0f172a')
    ax.axis('off')

    meta = data["metadata"]
    fin = data["financials"]

    # Colors
    c_blue = '#3b82f6'
    c_green = '#10b981'
    c_amber = '#f59e0b'
    c_red = '#ef4444'
    c_white = '#ffffff'
    c_muted = '#94a3b8'

    # Title & Header
    ax.text(0.05, 0.95, "MARUTHI — Business Snapshot", fontsize=22, fontweight='bold', color=c_white)
    ax.text(0.05, 0.92, f"Shop: {meta['shop_name']} ({meta['locality']})", fontsize=13, color=c_blue)
    ax.text(0.05, 0.895, f"Period: {meta['period_start']} to {meta['period_end']} ({meta['period_code']})", fontsize=11, color=c_muted)
    ax.text(0.05, 0.875, f"Generated: {meta['generated_at']}", fontsize=10, color=c_muted)

    ax.plot([0.05, 0.95], [0.86, 0.86], color='#334155', linewidth=1.5)

    # Key Metrics Cards (2x2 Grid)
    # Card 1: Revenue
    ax.text(0.05, 0.81, "REVENUE", fontsize=10, fontweight='bold', color=c_muted)
    ax.text(0.05, 0.77, f"Rs. {fin['revenue']:,.2f}", fontsize=20, fontweight='bold', color=c_white)

    # Card 2: Profit
    ax.text(0.52, 0.81, "GROSS PROFIT", fontsize=10, fontweight='bold', color=c_muted)
    ax.text(0.52, 0.77, f"Rs. {fin['profit']:,.2f}", fontsize=20, fontweight='bold', color=c_green)

    # Card 3: Gross Margin
    ax.text(0.05, 0.70, "GROSS MARGIN", fontsize=10, fontweight='bold', color=c_muted)
    ax.text(0.05, 0.66, f"{fin['margin']}%", fontsize=20, fontweight='bold', color=c_blue)

    # Card 4: Inventory Valuation
    ax.text(0.52, 0.70, "INVENTORY VALUE", fontsize=10, fontweight='bold', color=c_muted)
    ax.text(0.52, 0.66, f"Rs. {fin['inventory_value']:,.2f}", fontsize=20, fontweight='bold', color=c_amber)

    ax.plot([0.05, 0.95], [0.63, 0.63], color='#334155', linewidth=1.5)

    # Section: Top Sellers
    ax.text(0.05, 0.58, "TOP SELLER PRODUCTS", fontsize=13, fontweight='bold', color=c_white)
    y_pos = 0.54
    top_sellers = data["sales_performance"]["top_sellers"][:3]
    if top_sellers:
        for ts in top_sellers:
            ax.text(0.07, y_pos, f"• {ts['name']}", fontsize=11, color=c_white)
            ax.text(0.65, y_pos, f"{ts['units_sold']} units | Rs. {ts['revenue']:,.0f}", fontsize=11, color=c_muted)
            y_pos -= 0.035
    else:
        ax.text(0.07, y_pos, "No sales recorded during this period", fontsize=11, color=c_muted)
        y_pos -= 0.035

    ax.plot([0.05, 0.95], [y_pos-0.01, y_pos-0.01], color='#334155', linewidth=1.5)
    y_pos -= 0.05

    # Section: Tomorrow's Advisor Priority
    ax.text(0.05, y_pos, "TOMORROW'S PRIORITY ACTION", fontsize=13, fontweight='bold', color=c_amber)
    y_pos -= 0.035
    adv_recs = data["advisor"]["recommendations"]
    adv_title = adv_recs[0]["title"] if adv_recs else "Inventory healthy"
    adv_sum = adv_recs[0]["recommendation_summary"] if adv_recs else "No urgent reorders."

    ax.text(0.07, y_pos, f"Priority: {adv_title}", fontsize=11, fontweight='bold', color=c_white)
    y_pos -= 0.03
    ax.text(0.07, y_pos, adv_sum[:80] + "..." if len(adv_sum) > 80 else adv_sum, fontsize=10, color=c_muted)
    y_pos -= 0.045

    ax.plot([0.05, 0.95], [y_pos-0.01, y_pos-0.01], color='#334155', linewidth=1.5)
    y_pos -= 0.05

    # Section: Seasonal & Local Driver
    ax.text(0.05, y_pos, "SEASONAL & LOCAL SIGNALS", fontsize=13, fontweight='bold', color=c_green)
    y_pos -= 0.035
    sec_intel = data["seasonal_local"]
    ax.text(0.07, y_pos, f"Season: {sec_intel['current_season']} | Locality: {sec_intel['location_name']}", fontsize=11, color=c_white)
    y_pos -= 0.03
    fest_str = ", ".join(sec_intel['upcoming_festivals']) if sec_intel['upcoming_festivals'] else "Standard Retail Cycle"
    ax.text(0.07, y_pos, f"Upcoming Drivers: {fest_str}", fontsize=10, color=c_muted)
    y_pos -= 0.06

    # Footer
    ax.plot([0.05, 0.95], [0.06, 0.06], color='#334155', linewidth=1)
    ax.text(0.05, 0.03, "MARUTHI Verified Business Data | 100% Read-Only Safety Guarantee", fontsize=9, color=c_muted)

    buffer = io.BytesIO()
    plt.tight_layout()
    plt.savefig(buffer, format='png', facecolor=fig.get_facecolor(), bbox_inches='tight')
    plt.close(fig)
    buffer.seek(0)
    return buffer.getvalue()
